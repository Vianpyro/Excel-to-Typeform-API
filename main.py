#!/usr/bin/env python3
# -*- coding: utf8 -*-
import os
import subprocess
import tkinter as tk
import tkinter.filedialog as tkfiledialog
import tkinter.messagebox as tkmessagebox
import webbrowser

# 6NR1L7LJ2G2LmAb37yRkC3EUdeo7qfyX3wDnMXWxtod5
typeform_types = [
    'date', 'dropdown', 'email', 'file_upload', 
    'group', 'legal', 'long_text', 'matrix', 
    'multiple_choice', 'number', 'opinion_scale', 
    'payment', 'phone_number', 'picture_choice', 
    'ranking', 'rating', 'short_text', 'statement',
    'website', 'yes_no'
]

# Install requirements
subprocess.Popen('pip3 install -r requirements.txt')
import openpyxl
import typeform

###########################
# Workbook (Excel metadata)
###########################
class Workbook:
    def __init__(self, filename, read_only:bool=False, keep_vba:bool=False, data_only:bool=False, keep_links:bool=True):
        """
        Constructor for Workbook class.

        :filename: The name of the file to extract the informations from.
        :read_only: Optimised for reading, content cannot be edited.
        :keep_vba: Preseve vba content (this does NOT mean you can use it).
        :data_only: Controls whether cells with formulae have either the formula (default) or the value stored the last time Excel read the sheet.
        :keep_links: Whether links to external workbooks should be preserved (True by default).
        """
        self.workbook = openpyxl.load_workbook(str(filename), read_only, keep_vba, data_only, keep_links)
        self.sheets_count = len(self.workbook.sheetnames)
        self.content = [
            [
                [cell.value for cell in row]
                for row in self.workbook[sheet].rows
            ]
            for sheet in self.workbook.sheetnames
        ] if self.sheets_count > 1 else [
            [cell.value for cell in row]
            for row in self.workbook[self.workbook.sheetnames[0]].rows
        ]
        self.workbook.close()
        self.rows_titles = self.content[0] if self.sheets_count == 1 else [sheet[0] for sheet in self.content]


###########################
# Tkinter App
###########################
class Application(tk.Frame):
    def __init__(self, master: tk.Tk = None):
        """
        Constructor for the Graphic User Interface.

        :param master: The application's Tkinter root.
        """
        super().__init__(master)
        self.master = master
        self.widgets_width = 30
        self.master.title('Excel to Typeform')
        self.grid()
        self.create_widgets()

    def create_widgets(self) -> None:
        """
        Method creating and displaying each widget on the Graphic User Interface.
        """
        # Excel resource file
        self.excel_file_label = tk.Label(self.master, text='Excel resource file:')
        self.excel_file_label.grid(row=0, column=0, sticky='w')

        self.excel_file = tk.Entry(self.master, state='readonly', width=self.widgets_width)
        self.excel_file.grid(row=0, column=1, columnspan=2)

        self.excel_file_select_button = tk.Button(self.master, text='Select', padx=7, pady=2, command=self.select_excel_file)
        self.excel_file_select_button.grid(row=0, column=3)

        # Typeform title
        self.typeform_title_label = tk.Label(self.master, text='Typeform title:')
        self.typeform_title_label.grid(row=1, column=0, sticky='w')

        self.typeform_title = tk.Entry(self.master, width=self.widgets_width)
        self.typeform_title.grid(row=1, column=1, columnspan=2)
        
        # Typeform form type
        self.typeform_type_label = tk.Label(self.master, text='Typeform type:')
        self.typeform_type_label.grid(row=2, column=0, sticky='w')

        self.typeform_type_value = tk.StringVar(self.master)
        self.typeform_type_value.set('<!> auto <!>')
        self.typeform_type = tk.OptionMenu(self.master, self.typeform_type_value, *typeform_types)
        self.typeform_type.grid(row=2, column=1, columnspan=2)

        # Typeform token
        self.typeform_token_label = tk.Label(self.master, text='Typeform token:')
        self.typeform_token_label.grid(row=3, column=0, sticky='w')

        self.typeform_token = tk.Entry(self.master, width=self.widgets_width // 2, show='*')
        self.typeform_token.grid(row=3, column=1)
        self.typeform_token.focus()

        self.typeform_token_display_variable = tk.IntVar()
        self.typeform_token_display = tk.Checkbutton(self.master, text='Show token', variable=self.typeform_token_display_variable, command=self.show_token)
        self.typeform_token_display.grid(row=3, column=2)

        self.typeform_token_renew_button = tk.Button(
            self.master, text='Renew', padx=7, pady=2,
            command=lambda: webbrowser.open('https://admin.typeform.com/account#/section/tokens', new=1)
        )
        self.typeform_token_renew_button.grid(row=3, column=3)
        
        # Remove older forms
        self.remove_older_forms = tk.IntVar()
        self.remove_older_forms_check = tk.Checkbutton(self.master, text='Remove older Typeforms', variable=self.remove_older_forms)
        self.remove_older_forms_check.grid(row=self.master.grid_size()[1], column=0, columnspan=4, pady=10)

        # Start generation button
        self.generate_button = tk.Button(self.master, text='Generate Typeform', padx=10, pady=10, bg='green', font='bold', command=self.generate_api)

        if os.name == 'nt':
            self.generate_button.grid(row=self.master.grid_size()[1], column=0, columnspan=4, padx=self.widgets_width, pady=self.widgets_width * 3)
        else:
            self.generate_button.grid(row=self.master.grid_size()[1], column=2, columnspan=2, padx=self.widgets_width, pady=self.widgets_width * 3)
            
            # Close button
            self.close_button = tk.Button(text='Close', padx=10, pady=10, bg='red', font='bold', command=self.master.destroy)
            self.close_button.grid(row=self.master.grid_size()[1] - 1, column=0)

        # Console
        self.console_label = tk.Label(self.master, text='Console:')
        self.console_label.grid(row=self.master.grid_size()[1], column=0, sticky='w')

        self.console = tk.Text(self.master, state=tk.DISABLED, width=self.widgets_width + 10, height=7)
        self.console.grid(row=self.master.grid_size()[1], column=0, columnspan=4)

    def show_token(self) -> None:
        """
        Method dealing with the visibility of the token.
        """
        if self.typeform_token_display_variable.get():
            self.typeform_token['show'] = ''
        else:
            self.typeform_token['show'] = '*'

    def select_excel_file(self) -> None:
        """
        Method for choosing the Excel file to be used
        """
        # Ask user to select the file to open
        filename = tkfiledialog.askopenfilename(initialdir=os.path.abspath(os.getcwd()), filetypes=[('Excel files', '.xlsx .xlsm .xltx .xltm .xls')])

        # Display the file name
        self.excel_file['state'] = 'normal'
        self.excel_file.delete(0, tk.END)
        self.excel_file.insert(tk.END, filename.split('/')[-1])
        self.excel_file['state'] = 'readonly'
        self.console_log(f'Opened {filename}' if filename != '' else 'No file selected')

        # Auto generate Typeform title
        self.typeform_title.delete(0, tk.END)
        self.typeform_title.insert(tk.END, filename.split('/')[-1].split('.')[0])

    def console_log(self, message:str='<message>') -> None:
        """
        Method to log a message in the console box of the Graphic User Interface.
        """
        self.console['state'] = 'normal'
        self.console.insert(tk.END, f'> {message}.\n')
        self.console.see('end')
        self.console['state'] = tk.DISABLED

    def generate_api(self) -> None:
        """
        Method to make the request to the Typeform API.
        """
        if self.verify_widgets():
            self.console_log('Generating Typeform..')

            wb = Workbook(filename=self.excel_file.get(), read_only=True)
            my_typeform = typeform.Typeform(self.typeform_token.get())
            form = my_typeform.forms
            
            # Delete older Typeforms ?
            if self.remove_older_forms.get():
                if tkmessagebox.askyesno(
                    'Delete older form(s) confirmation',
                    f'You are about to delete {len(form.list()["items"])} form(s), are you sure you want to continue?'
                ):
                    for e in form.list()['items']:
                        form.delete(e['id'])
                else:
                    self.console_log('Canceled Typeform deletion')

            fields = [
                {
                    "title": wb.content[i + 1][1],
                    "ref": f"Question-{wb.content[i + 1][0]}",
                    "type": str(self.typeform_type_value.get() if self.typeform_type_value.get() in typeform_types else 'multiple_choice'),
                    "properties": {
                        "randomize": False,
                        "allow_multiple_selection": False,
                        "allow_other_choice": False,
                        "vertical_alignment": True,
                        "choices": [
                            {
                                "label": label,
                                "ref": "test"+label
                            }
                            for label in wb.content[i + 1][3].split('/')
                        ]
                    },
                    "validations": {"required": True}
                }
                for i in range(len(wb.content) - 1)
            ]
            
            # Create a new form
            new_form = form.create({'title': self.typeform_title.get(), 'fields': fields})
            self.console_log(f'New form created with ID: {new_form["id"]}')
            self.console_log('You can now close this window')
        else:
            self.console_log('Error - At least one field is not correct')

    def verify_widgets(self) -> bool:
        """
        Method to check if a request can be made to the API or if the user still has to fill in at least one field.
        
        :return: Wether a proper request can be made to the API.
        """
        return bool(
            self.excel_file.get() != ''
            and self.typeform_title.get() != ''
            and len(self.typeform_token.get()) == 44
        )

root = tk.Tk()
app = Application(master=root)
app.mainloop()
